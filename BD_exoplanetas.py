#Functions

global Sector
global Time
global Flux
global Writexl
global Error
global ID
global Quantity
global lc

def ligthcurve(Telescope,Method,Planet,Quantity,i):
    
    Flux = 0
    Time = 0
    Error = 0
    
    if Telescope == Satellite1 and Method == Discmethod:
        
        try:
            
            lcf = search_lightcurvefile(Planet,mission='Kepler').download_all(download_dir = MAST)
                
            lc = lcf.PDCSAP_FLUX.stitch().remove_nans().remove_outliers().flatten(window_length = 1001, polyorder = 3, niters=5)
        
            Time = lc.time
            Flux = lc.flux
            Error = lc.flux_err

            Writexl = hoja.cell(row = i, column = 25) 
            Writexl.value = "Found"
            
            file = createfile(Planet,Time,Flux,Error,Quantity)
            
            BLS(lc,file,i)
            
        
        except KeyError:
            
            RA = hoja.cell(row = i, column = 16) 
            Dec = hoja.cell(row = i, column = 18) 
            
            Coordenada = str(RA.value) +" " + str(Dec.value)
            
            lcf = search_lightcurvefile(Coordenada,mission='Kepler').download_all(download_dir = MAST)
                
            lc = lcf.PDCSAP_FLUX.stitch().remove_nans().remove_outliers().flatten(window_length = 1001, polyorder = 3, niters=5)
        
            Time = lc.time
            Flux = lc.flux
            Error = lc.flux_err
            
            Writexl = hoja.cell(row = i, column = 25) 
            Writexl.value = "Found"
            
            file = createfile(Planet,Time,Flux,Error,Quantity)
            
            BLS(lc,file,i)
            
            
        except Exception as e:
            print(str(e))

#-------------------------------------------------------------------------
            
    elif Telescope == Satellite2 and Method == Discmethod:
        
        try:
            
            lcf = search_lightcurvefile(Planet,mission='K2').download_all(download_dir = MAST)
                
            lc = lcf.PDCSAP_FLUX.stitch().remove_nans().remove_outliers().flatten(window_length = 1001, polyorder = 3, niters=5)
        
            Time = lc.time
            Flux = lc.flux
            Error = lc.flux_err
    
            Writexl = hoja.cell(row = i, column = 25) 
            Writexl.value = "Found"
            
            createfile(Planet,Time,Flux,Error,Quantity)
            
            file = createfile(Planet,Time,Flux,Error,Quantity)
            
            BLS(lc,file,i)
            
        except KeyError:
            
            RA = hoja.cell(row = i, column = 16) 
            Dec = hoja.cell(row = i, column = 18) 
            
            Coordenada = str(RA.value) +" " + str(Dec.value)
            
            lcf = search_lightcurvefile(Coordenada,mission='K2').download_all(download_dir = MAST)
                
            lc = lcf.PDCSAP_FLUX.stitch().remove_nans().remove_outliers().flatten(window_length = 1001, polyorder = 3, niters=5)
        
            Time = lc.time
            Flux = lc.flux
            Error = lc.flux_err
            
            Writexl = hoja.cell(row = i, column = 25) 
            Writexl.value = "Found"
            
            file = createfile(Planet,Time,Flux,Error,Quantity)
            
            BLS(lc,file,i)

        except Exception as e:
            print(str(e))
        
#-------------------------------------------------------------------------------

    elif Telescope == Satellite3 and Method == Discmethod:
        
        try:
            
            lcf = search_lightcurvefile(Planet,mission='TESS').download_all(download_dir = MAST)
                
            lc = lcf.PDCSAP_FLUX.stitch().remove_nans().remove_outliers().flatten(window_length = 1001, polyorder = 3, niters=5)
        
            Time = lc.time
            Flux = lc.flux
            Error = lc.flux_err

            Writexl = hoja.cell(row = i, column = 25) 
            Writexl.value = "Found"
            
            file = createfile(Planet,Time,Flux,Error,Quantity)
            
            BLS(lc,file,i)
        
        except KeyError:
            
            RA = hoja.cell(row = i, column = 16) 
            Dec = hoja.cell(row = i, column = 18) 
            
            Coordenada = str(RA.value) +" " + str(Dec.value)
            
            lcf = search_lightcurvefile(Coordenada,mission='TESS').download_all(download_dir = MAST)
                
            lc = lcf.PDCSAP_FLUX.stitch().remove_nans().remove_outliers().flatten(window_length = 1001, polyorder = 3, niters=5)
            
            Time = lc.time
            Flux = lc.flux
            Error = lc.flux_err
            
            Writexl = hoja.cell(row = i, column = 25) 
            Writexl.value = "Found"
            
            file = createfile(Planet,Time,Flux,Error,Quantity)
            
            BLS(lc,file,i)

        except Exception as e:
            print(str(e))
            
#--------------------------------------------------------------------------------

    else:
        Writexl = hoja.cell(row = i, column = 25) 
        Writexl.value = "No search"
        
#--------------------------------------------------------------------------------

def createfile(Planetfile,Timefile,Fluxfile,Errorfile,Quantity):
    
    print(" ")
    print(i)
    print(Planet.value)
    
    if Quantity == 1:
        
        save_path = path + "\Planet_1"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
    elif Quantity == 2:
        
        save_path = path + "\Planet_2"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
    elif Quantity == 3:
        
        save_path = path + "\Planet_3"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
    elif Quantity == 4:
        
        save_path = path + "\Planet_4"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
    elif Quantity == 5:
        
        save_path = path + "\Planet_5"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
    elif Quantity == 6:
        
        save_path = path + "\Planet_6"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
        
    elif Quantity == 7:
        
        save_path = path + "\Planet_7"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
    
    elif Quantity == 8:
        
        save_path = path + "\Planet_8"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1) 
        
        return file
        
    else:

        save_path = path + "\Other"
        
        file = save_path+'\\'+Planetfile+".txt" 
        np.savetxt(file, (Fluxfile, Timefile, Errorfile),delimiter = ',', newline = '\n\n')
        
        Writexl = hoja.cell(row = i, column = 26) 
        Writexl.value = "File created"
        
        doc.save(path + "\\" + excel)
        
        time.sleep(1)
        
        return file
    
#-------------------------------------------------------------------------------------------------

def BLS(lcs,file,i):
    
    duration = 0.25
    
    min_period = np.max([np.median(np.diff(lcs.time)) * 4,np.max(duration) + np.median(np.diff(lcs.time))])
    max_period = (np.max(lcs.time) - np.min(lcs.time)) / 3.
    
        
    df_min = (max_period-min_period)/(1e5*(max_period*min_period))
    freq_factor = (df_min*((np.max(lcs.time) - np.min(lcs.time))**2))/duration
    
    if max_period > 365:
        max_period = 365
    
    if min_period < 1:
        min_period = 1
        
    pgs = lcs.to_periodogram(method = 'bls', frequency_factor = freq_factor, minimum_period = min_period, maximum_period = max_period) 
    
    period_pgs = pgs.period_at_max_power.value
    t_transit = pgs.transit_time_at_max_power
    
    power = pgs.power.value
    power[power < 0] = 0
    period = pgs.period.value
    depth = pgs.depth_at_max_power

    
    flux_fold = lcs.fold(period_pgs,t0 = t_transit).flux
    phase_fold = lcs.fold(period_pgs,t0 = t_transit).phase
    
    power = ["%.7f" % number for number in power]
    power = ','.join(power)
    period = ["%.7f" % number for number in period]
    period = ','.join(period)
    flux_fold = ["%.7f" % number for number in flux_fold]
    flux_fold = ','.join(flux_fold)
    phase_fold = ["%.7f" % number for number in phase_fold]
    phase_fold = ','.join(phase_fold)
    
    with open(file,'a') as f:

        f.write(power + "\n\n")
        f.write(period  + "\n\n")
        f.write(flux_fold  + "\n\n")
        f.write(phase_fold  + "\n\n")
        f.write(str(period_pgs)  + "\n\n")
        f.write(str(t_transit)  + "\n\n")
        f.write(str(depth))
        
    Writexl = hoja.cell(row = i, column = 27)
    Writexl.value = depth
        
    doc.save(path + "\\" + excel)
    
    time.sleep(0.5)
    
#------------------------------MAIN CODE-------------------------------------------

from lightkurve import search_lightcurvefile
import openpyxl
import time
import numpy as np
import os


path = r"C:\Users\HP\Desktop\Prueba_BD"
excel = "planets_2020.05.16_11.23.37.xlsx"
sheet = "planets_2020.05.16_11.23.37"

for i in range(1,9):
    folder = path + "\Planet_" + str(i)
    os.mkdir(folder)

os.mkdir(path + "\Other")

MAST = os.mkdir(path + "\MAST")

doc = openpyxl.load_workbook(path + "\\" + excel)

hoja = doc[sheet]

Satellite1 = 'Kepler'
Satellite2 = 'K2'
Satellite3 = 'Transiting Exoplanet Survey Satellite (TESS)'
Discmethod = 'Transit'

Writexl = hoja.cell(row = 29, column = 25)
Writexl.value = "Light Curve"
Writexl = hoja.cell(row = 29, column = 26)
Writexl.value = "File Status"
Writexl = hoja.cell(row = 29, column = 27)
Writexl.value = "Depth"
        

for i in range(1262,1263):
    Planet = hoja.cell(row = i, column = 4) 
    Telescope = hoja.cell(row = i, column = 23)
    Method = hoja.cell(row = i, column = 5)
    Quantity = hoja.cell(row = i, column = 6)
    
    try:
        
        ligthcurve(Telescope.value,Method.value,Planet.value,Quantity.value,i)

    except Exception as exc:
        print(exc)
        Writexl = hoja.cell(row = i, column = 25) 
        Writexl.value = "Not found"

    
doc.save(path + "\\" + excel)
